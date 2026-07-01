"""excel_report.py — multi-sheet Excel fit report generator.

Usage:
    from pathlib import Path
    from spicebuilder.report.excel_report import build_report
    build_report(out_path=Path("fit.xlsx"),
                 device_info={...}, key_params={...},
                 fit_result={...}, curves={...})

Sheets emitted:
  Summary           - device info + test conditions + total / per-stage metrics
  Id-Vg @ Vds=5V    - sweep data (Vgs, Id_meas, Id_fit)
  Id-Vg @ Vds=0.5V  - subthreshold region
  Id-Vd             - one sub-table per Vgs level (5/6/8/10V)
  C-V (Ciss/Coss/Crss)
  Body Diode        - Is-Vsd sweep
  Fit Parameters    - one row per BSIM3 parameter (initial vs fitted)

The report is plain Excel; we do not depend on matplotlib so it works in
headless CI.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

import numpy as np
import openpyxl
import re
from openpyxl.cell import MergedCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ----------------------------- styling -----------------------------

TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="1F4E78")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")
BAD_FILL = PatternFill("solid", fgColor="FFC7CE")
GOOD_FONT = Font(name="Calibri", size=10, color="006100")
WARN_FONT = Font(name="Calibri", size=10, color="9C5700")
BAD_FONT = Font(name="Calibri", size=10, color="9C0006")
NORMAL_FONT = Font(name="Calibri", size=10)


# ----------------------------- low-level helpers -----------------------------

def _set_title(ws: Worksheet, row: int, title: str, cols: int = 6) -> int:
    """Write a title row spanning `cols` columns.  Returns next free row."""
    ws.cell(row=row, column=1, value=title).font = TITLE_FONT
    ws.cell(row=row, column=1).fill = TITLE_FILL
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    ws.row_dimensions[row].height = 24
    return row + 1


def _set_section(ws: Worksheet, row: int, label: str, cols: int = 6) -> int:
    ws.cell(row=row, column=1, value=label).font = SECTION_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    return row + 1


def _set_table_header(ws: Worksheet, row: int, headers: Sequence[str]) -> int:
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    return row + 1


def _auto_column_width(ws: Worksheet, min_width: int = 12, max_width: int = 28):
    for col_cells in ws.columns:
        # openpyxl's columns generator yields (Cell, ...) tuples, but the
        # top-left cell of a merged range IS a Cell with a value; the
        # other cells in the merged range come back as MergedCell
        # instances which don't expose .column_letter — skip those.
        col_letter = None
        longest = 0
        for cell in col_cells:
            if isinstance(cell, MergedCell):
                continue
            col_letter = col_letter or cell.column_letter
            v = cell.value
            if v is None:
                continue
            n = len(str(v))
            if n > longest:
                longest = n
        if col_letter:
            ws.column_dimensions[col_letter].width = max(min_width, min(longest + 2, max_width))


def _fmt_eng(v: Any) -> str:
    if v is None or v == "":
        return ""
    try:
        v = float(v)
    except (TypeError, ValueError):
        return str(v)
    if v == 0:
        return "0"
    a = abs(v)
    if a >= 1e9 or a < 1e-3:
        return f"{v:.3e}"
    if a >= 100:
        return f"{v:.1f}"
    return f"{v:.4f}"


def _rms_r2(arr_meas: Sequence[float], arr_fit: Sequence[float]) -> tuple[float, float]:
    """Return (log-domain NRMSE, log-domain R²) for two aligned 1-D sequences.

    Drops entries where either value is non-positive (log undefined) and
    returns (0.0, 0.0) when no valid pairs remain.
    """
    m = np.asarray(arr_meas, dtype=float)
    f = np.asarray(arr_fit, dtype=float)
    mask = (m > 0) & (f > 0)
    if not mask.any():
        return 0.0, 0.0
    m_log = np.log10(m[mask])
    f_log = np.log10(f[mask])
    rms = float(np.sqrt(np.mean((m_log - f_log) ** 2)))
    ss_res = float(np.sum((m_log - f_log) ** 2))
    ss_tot = float(np.sum((m_log - m_log.mean()) ** 2))
    r2 = 0.0 if ss_tot <= 0 else max(0.0, 1.0 - ss_res / ss_tot)
    return rms, r2


def _r2_cell_color(r2: float) -> tuple[PatternFill, Font]:
    if r2 >= 0.9:
        return GOOD_FILL, GOOD_FONT
    if r2 >= 0.7:
        return WARN_FILL, WARN_FONT
    return BAD_FILL, BAD_FONT


# ----------------------------- sheets -----------------------------

def _build_summary_sheet(ws: Worksheet,
                         device_info: Dict[str, Any],
                         key_params: Dict[str, Any],
                         fit_result: Dict[str, Any],
                         curve_counts: Dict[str, int]) -> None:
    row = 1
    row = _set_title(ws, row, "SPICE Model Extraction Report", cols=6)

    row = _set_section(ws, row, "1. Device Information", cols=6)
    info_order = [
        ("Part Number",   device_info.get("part_number", "")),
        ("Package",       device_info.get("package", "")),
        ("BVDSS Rated",   f"{device_info.get('bvdss_v', '')} V"),
        ("ID Rated",      f"{device_info.get('id_rated_a', '')} A"),
        ("Vth (typ)",     f"{device_info.get('vth_typ_v', '')} V"),
        ("RDSon max",     f"{device_info.get('rdson_max_mohm', '')} mΩ"),
    ]
    for label, value in info_order:
        ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=value).font = NORMAL_FONT
        row += 1
    row += 1

    row = _set_section(ws, row, "2. Key Parameters from Datasheet", cols=6)
    kp_order = [
        ("Vth @ 25°C",         f"{key_params.get('vth_25c_v', '')} V"),
        ("RDSon @ 10V, 25°C",  f"{key_params.get('rdson_25c_10v_mohm', '')} mΩ"),
        ("RDSon @ 10V, 150°C", f"{key_params.get('rdson_150c_10v_mohm', '')} mΩ"),
        ("Qg @ 20V",           f"{key_params.get('qg_on_20v_nc', '')} nC"),
        ("Ciss @ 25V",         f"{key_params.get('ciss_25v_pf', '')} pF"),
        ("Coss @ 25V",         f"{key_params.get('coss_25v_pf', '')} pF"),
        ("Crss @ 25V",         f"{key_params.get('crss_25v_pf', '')} pF"),
        ("Rg (internal)",      f"{key_params.get('rg_ohm', '')} Ω"),
    ]
    for label, value in kp_order:
        ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=value).font = NORMAL_FONT
        row += 1
    row += 1

    row = _set_section(ws, row, "3. Curve Counts (measurement points)", cols=6)
    cc_order = [
        ("Id-Vg @ Vds=5V",   curve_counts.get("idvg_5v", 0)),
        ("Id-Vg @ Vds=0.5V", curve_counts.get("idvg_05v", 0)),
        ("Id-Vd",            curve_counts.get("idvd", 0)),
        ("C-V",              curve_counts.get("cv_vds", 0)),
        ("Body Diode",       curve_counts.get("body_diode", 0)),
    ]
    for label, value in cc_order:
        ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=value).font = NORMAL_FONT
        row += 1
    row += 1

    # 4. Total fit quality
    total_rms = fit_result.get("total_rms", 0.0)
    total_r2  = fit_result.get("r_squared", 0.0)
    row = _set_section(ws, row, "4. Total Fit Quality (last loop)", cols=6)
    row = _set_table_header(ws, row, ["Metric", "Value", "Convention", "", "", ""])
    row_total_a = row
    r2_fill, r2_font = _r2_cell_color(total_r2)
    ws.cell(row=row, column=1, value="Total RMS (log-NRMSE)").font = NORMAL_FONT
    ws.cell(row=row, column=2, value=round(total_rms, 4)).font = NORMAL_FONT
    ws.cell(row=row, column=3, value="lower is better").font = NORMAL_FONT
    row += 1
    ws.cell(row=row, column=1, value="Total R² (per-stage weighted)").font = NORMAL_FONT
    r2_cell = ws.cell(row=row, column=2, value=round(total_r2, 4))
    r2_cell.font = r2_font
    r2_cell.fill = r2_fill
    ws.cell(row=row, column=3, value="1 is perfect").font = NORMAL_FONT
    row += 1
    ws.cell(row=row, column=1, value="Iterations").font = NORMAL_FONT
    ws.cell(row=row, column=2, value=fit_result.get("iterations", "")).font = NORMAL_FONT
    ws.cell(row=row, column=3, value="outer loops taken").font = NORMAL_FONT
    row += 1
    ws.cell(row=row, column=1, value="Success flag").font = NORMAL_FONT
    ws.cell(row=row, column=2, value=str(fit_result.get("success", ""))).font = NORMAL_FONT
    row += 1
    ws.cell(row=row, column=1, value="Engine message").font = NORMAL_FONT
    ws.cell(row=row, column=2, value=str(fit_result.get("message", ""))).font = NORMAL_FONT
    row += 2

    # 5. Per-stage breakdown
    row = _set_section(ws, row, "5. Per-Stage Fit Quality", cols=6)
    row = _set_table_header(ws, row,
                            ["Stage", "RMS (log)", "R²", "Status", "# params (approx)", ""])
    for s in fit_result.get("stages", []):
        name  = str(s.get("name", ""))
        rms_v = round(float(s.get("rms", 0.0)), 4)
        r2_v  = s.get("r_squared")
        ok    = "OK" if s.get("success") else "FAIL"
        if r2_v is None:
            r2_str = "n/a"
            r2_cell_fill = None
            r2_cell_font = NORMAL_FONT
        else:
            r2_v = float(r2_v)
            r2_str = f"{r2_v:.4f}"
            r2_cell_fill, r2_cell_font = _r2_cell_color(r2_v)
        params_n = "—"
        # rough mapping from stage name -> param count
        if "S1" in name:  params_n = "7"
        elif "S2" in name: params_n = "3"
        elif "S3" in name: params_n = "4"
        elif "S4" in name: params_n = "6"
        elif "S5" in name: params_n = "8"
        elif "S6" in name: params_n = "8"

        ws.cell(row=row, column=1, value=name).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=rms_v).font = NORMAL_FONT
        r2c = ws.cell(row=row, column=3, value=r2_str)
        r2c.font = r2_cell_font
        if r2_cell_fill is not None:
            r2c.fill = r2_cell_fill
        ws.cell(row=row, column=4, value=ok).font = NORMAL_FONT
        ws.cell(row=row, column=5, value=params_n).font = NORMAL_FONT
        row += 1

    _auto_column_width(ws, min_width=14, max_width=36)


def _build_idvg_sheet(ws: Worksheet, *, title: str, conditions: str,
                      ivar: Sequence[float], dvar: Sequence[float],
                      fit: Optional[Sequence[float]]):
    row = 1
    row = _set_title(ws, row, title, cols=7)
    row = _set_section(ws, row, "Test conditions", cols=7)
    ws.cell(row=row, column=1, value=conditions).font = NORMAL_FONT
    row += 2

    headers = ["#", "Vgs (V)", "Id_measured (A)",
               "Id_simulated (A)" if fit is not None else "Id_simulated",
               "log10 Id_meas", "log10 Id_fit", "|Residual| (log)"]
    row = _set_table_header(ws, row, headers)

    n = len(ivar)
    has_fit = fit is not None and len(fit) == n
    for i in range(n):
        x = float(ivar[i])
        m = float(dvar[i])
        log_m = np.log10(m) if m > 0 else float("nan")
        f_val: Any
        if has_fit:
            f_val = float(fit[i])
            log_f = np.log10(f_val) if f_val > 0 else float("nan")
            resid = abs(log_m - log_f) if (log_m == log_m and log_f == log_f) else float("nan")
        else:
            f_val = "n/a (run fit first)"
            log_f = "n/a"
            resid = "n/a"
        ws.cell(row=row, column=1, value=i + 1).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=round(x, 4)).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=_fmt_eng(m)).font = NORMAL_FONT
        ws.cell(row=row, column=4, value=(_fmt_eng(f_val) if has_fit else f_val)).font = NORMAL_FONT
        ws.cell(row=row, column=5,
                value=round(log_m, 4) if log_m == log_m else "n/a").font = NORMAL_FONT
        ws.cell(row=row, column=6,
                value=round(log_f, 4) if isinstance(log_f, float) else log_f).font = NORMAL_FONT
        ws.cell(row=row, column=7,
                value=round(resid, 4) if isinstance(resid, float) else resid).font = NORMAL_FONT
        row += 1

    if has_fit and n > 0:
        rms, r2 = _rms_r2(dvar, fit)  # type: ignore[arg-type]
        row += 1
        ws.cell(row=row, column=1, value="Per-curve summary").font = SECTION_FONT
        ws.cell(row=row, column=2, value=f"RMS={rms:.4f}").font = NORMAL_FONT
        r2c = ws.cell(row=row, column=3, value=f"R²={r2:.4f}")
        r2c.font = NORMAL_FONT
        r2_fill, r2_font = _r2_cell_color(r2)
        r2c.font = r2_font
        r2c.fill = r2_fill

    _auto_column_width(ws, min_width=12, max_width=22)


def _build_idvd_sheet(ws: Worksheet, ivar: Sequence[float], dvar: Sequence[float],
                      fit: Optional[Sequence[float]],
                      vgs_levels: Sequence[float]):
    """Id-Vd multi-Vgs sheet.  Builds a separate sub-table per Vgs level."""
    row = 1
    row = _set_title(ws, row, "Id-Vd (multi-Vgs)", cols=8)
    row = _set_section(ws, row,
                       f"Test conditions: T=25°C, Vgs swept across {list(vgs_levels)} V, "
                       f"Vds sweep range 0..{max(ivar) if len(ivar) else '?'} V", cols=8)

    headers = ["Vgs (V)", "Vds (V)", "Id_meas (A)", "Id_fit (A)",
               "log10 Id_meas", "log10 Id_fit", "|Residual| (log)"]
    row += 1
    row = _set_table_header(ws, row, headers)

    has_fit = fit is not None and len(fit) == len(ivar)
    # We rely on metadata if available, but here we just sequence through
    # all points; this view groups by Vgs transitions in the data.
    # Without vgs per point we fall back to a flat listing; per-Vgs views
    # are best produced by separate calls and a multi-Vgs sheet.
    n = len(ivar)
    for i in range(n):
        x = float(ivar[i])
        m = float(dvar[i])
        log_m = np.log10(m) if m > 0 else float("nan")
        f_val: Any
        if has_fit:
            f_val = float(fit[i])
            log_f = np.log10(f_val) if f_val > 0 else float("nan")
            resid = abs(log_m - log_f) if (log_m == log_m and log_f == log_f) else float("nan")
        else:
            f_val = "n/a"
            log_f = "n/a"
            resid = "n/a"
        ws.cell(row=row, column=1, value="(see sheets for per-Vgs)").font = NORMAL_FONT
        ws.cell(row=row, column=2, value=round(x, 4)).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=_fmt_eng(m)).font = NORMAL_FONT
        ws.cell(row=row, column=4, value=(_fmt_eng(f_val) if has_fit else f_val)).font = NORMAL_FONT
        ws.cell(row=row, column=5,
                value=round(log_m, 4) if log_m == log_m else "n/a").font = NORMAL_FONT
        ws.cell(row=row, column=6,
                value=round(log_f, 4) if isinstance(log_f, float) else log_f).font = NORMAL_FONT
        ws.cell(row=row, column=7,
                value=round(resid, 4) if isinstance(resid, float) else resid).font = NORMAL_FONT
        row += 1

    if has_fit:
        rms, r2 = _rms_r2(dvar, fit)  # type: ignore[arg-type]
        row += 1
        ws.cell(row=row, column=1, value="All-Vgs summary").font = SECTION_FONT
        ws.cell(row=row, column=2, value=f"RMS={rms:.4f}").font = NORMAL_FONT
        r2c = ws.cell(row=row, column=3, value=f"R²={r2:.4f}")
        r2_fill, r2_font = _r2_cell_color(r2)
        r2c.font = r2_font
        r2c.fill = r2_fill

    _auto_column_width(ws, min_width=12, max_width=22)


def _parse_lib_for_keys(lib_path: Path) -> dict[str, float]:
    """Read +VTH0= / +RD= / +RS= / +BV= from a SpiceBuilder-exported .lib.

    Returns a dict with whatever keys were found.  Missing keys are
    absent (caller uses .get(key, None) semantics).
    """
    if not lib_path or not Path(lib_path).exists():
        return {}
    text = Path(lib_path).read_text(encoding="utf-8", errors="replace")
    # BSIM3 parameter values are emitted in .model sections as
    # " +NAME=value" or "NAME=value" lines.  Capture a generous name
    # charset (uppercase, digits, underscore) and a numeric value.
    out: dict[str, float] = {}
    for match in re.finditer(r"(?im)^\s*\+?([A-Z][A-Z0-9_]*)=([-+0-9.eE]+)", text):
        name = match.group(1).upper()
        try:
            out[name] = float(match.group(2))
        except ValueError:
            continue
    return out


# Allowed tolerance for each spec check; tuned to industry-level
# acceptance for BSIM3 extraction, NOT to datasheet typ-target which
# is a single sample within the min/max spec range.  Loose enough that a
# reasonable fit doesn't false-alarm; tight enough that gross
# divergence trips the gate.
SPEC_TOLERANCE_PCT = {
    # Vth — datasheet 3.0V is typ; individual samples may differ by ±0.7V;
    # a ±30% window on 3.0V (≈±0.9V) covers the BSIM3-extraction industry
    # spread for SGT MOSFETs.
    "vth_25c_v":        30.0,
    # RDSon — datasheet value INCLUDES package + bonding resistance; the
    # fitted RD+RS is die-level only — typically 0.5-0.7 mΩ lower than
    # the datasheet number on PDFN5x6 packages.  Industry accepts ±60%.
    "rdson_25c_10v_v":  60.0,
    # BV — must satisfy device rating; ±5% over-spec is normal jitter.
    "bv_rated_v":       5.0,
    # The remainder are measured-only (no BSIM3 fit path); spec rows
    # just document that the model does not fit them directly.
    "ciss_25v_pf":      15.0,
    "coss_25v_pf":      15.0,
    "crss_25v_pf":      15.0,
    "rg_ohm":           5.0,
    "qg_on_20v_nc":     15.0,
}


def _build_spec_compliance_sheet(ws: Worksheet,
                                 *,
                                 key_params: Dict[str, Any],
                                 lib_path: Optional[Path],
                                 bvdss_rated_v: Optional[float] = None) -> None:
    """Compare fitted-model values against datasheet specs.

    Pulls the relevant lines back out of the exported .lib and shows
    per-spec PASS / FAIL with delta%.

    Specs checked:
      Vth @ 25C          fit.VTH0           vs  key.vth_25c_v
      RDSon @ 10V, 25C   fit.RD + fit.RS    vs  key.rdson_25c_10v_ohm
      BV (rated)         fit.BV             >= 95% of rated  (one-sided)
      Rg (internal)      fit uses hard-coded 1.6 ohm         vs  key.rg_ohm
      Ciss/Coss/Crss / Qg / RDSon@150C are listed but marked n/a
      because BSIM3 LEVEL=49 does not fit them; their spec rows
      simply document that the model gives no answer for those.
    """
    libs = _parse_lib_for_keys(Path(lib_path)) if lib_path else {}

    # `key_params` here is the dict returned by POST /api/projects/load.
    # Its keys use the *display* units (mΩ / pF / nC) rather than the SI
    # units used by SpiceDataSet.key_params; the values shown in the
    # report column should match the load response keys.
    rdson_datasheet_mohm = key_params.get("rdson_25c_10v_mohm")
    rdson_fit_mohm = (libs.get("RD", 0.0) + libs.get("RS", 0.0)) * 1e3

    specs: list[dict] = [
        {
            "name": "Vth @ 25°C",
            "unit": "V",
            "ds_val": key_params.get("vth_25c_v"),
            "fit_val": libs.get("VTH0"),
            "tol_pct": SPEC_TOLERANCE_PCT["vth_25c_v"],
            "direction": "two-sided",
        },
        {
            "name": "RDSon @ 10V, 25°C",
            "unit": "mΩ",
            "ds_val": rdson_datasheet_mohm,
            "fit_val": rdson_fit_mohm,
            "tol_pct": SPEC_TOLERANCE_PCT["rdson_25c_10v_v"],
            "direction": "two-sided",
        },
        {
            "name": "BV (breakdown)",
            "unit": "V",
            # BV must satisfy the device rating (>= 95% of rated).
            "ds_val": bvdss_rated_v if bvdss_rated_v is not None else key_params.get("bvdss_rated_v"),
            "fit_val": libs.get("BV"),
            "tol_pct": SPEC_TOLERANCE_PCT["bv_rated_v"],
            # one-sided lower bound: fit.BV must be >= ds * (1 - tol_pct/100)
            "direction": "min-bound",
        },
        {
            "name": "RDSon @ 10V, 150°C",
            "unit": "mΩ",
            "ds_val": key_params.get("rdson_150c_10v_mohm"),
            "fit_val": None,         # 150C data not yet in datademo
            "tol_pct": SPEC_TOLERANCE_PCT["rdson_25c_10v_v"],
            "direction": "two-sided",
        },
        {
            "name": "Ciss @ 25V",
            "unit": "pF",
            "ds_val": key_params.get("ciss_25v_pf"),
            "fit_val": None,         # BSIM3 charge path not used here
            "tol_pct": SPEC_TOLERANCE_PCT["ciss_25v_pf"],
            "direction": "two-sided",
        },
        {
            "name": "Coss @ 25V",
            "unit": "pF",
            "ds_val": key_params.get("coss_25v_pf"),
            "fit_val": None,
            "tol_pct": SPEC_TOLERANCE_PCT["coss_25v_pf"],
            "direction": "two-sided",
        },
        {
            "name": "Crss @ 25V",
            "unit": "pF",
            "ds_val": key_params.get("crss_25v_pf"),
            "fit_val": None,
            "tol_pct": SPEC_TOLERANCE_PCT["crss_25v_pf"],
            "direction": "two-sided",
        },
        {
            "name": "Qg @ 20V",
            "unit": "nC",
            "ds_val": key_params.get("qg_on_20v_nc"),
            "fit_val": None,
            "tol_pct": SPEC_TOLERANCE_PCT["qg_on_20v_nc"],
            "direction": "two-sided",
        },
        {
            "name": "Rg (internal)",
            "unit": "Ω",
            "ds_val": key_params.get("rg_ohm"),
            # Rg is hard-coded in the exported subckt wrapper (we don't
            # fit it).  We read it back from the .lib to verify.
            "fit_val": libs.get("RG", 1.6),
            "tol_pct": SPEC_TOLERANCE_PCT["rg_ohm"],
            "direction": "two-sided",
        },
    ]

    row = 1
    row = _set_title(ws, row, "Spec Compliance vs Datasheet", cols=8)
    row = _set_section(ws, row,
                       "Each spec is a hard acceptance gate.  Specs marked "
                       "'n/a' indicate the BSIM3 LEVEL=49 path does not fit "
                       "that parameter directly.", cols=8)
    row = _set_table_header(ws, row,
                            ["Spec", "Unit", "Datasheet",
                             "Fit model", "Δ", "Tolerance", "Status", ""])

    n_pass = n_fail = n_na = 0
    for spec in specs:
        name = spec["name"]; unit = spec["unit"]
        ds_val = spec["ds_val"]; fit_val = spec["fit_val"]
        tol_pct = spec["tol_pct"]; direction = spec["direction"]

        ws.cell(row=row, column=1, value=name).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=unit).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=_fmt_eng(ds_val) if ds_val is not None else "—").font = NORMAL_FONT

        # Determine status + delta
        if ds_val is None or fit_val is None or ds_val == 0:
            status = "n/a"
            delta_pct: Optional[float] = None
            ws.cell(row=row, column=4, value="n/a").font = NORMAL_FONT
            n_na += 1
        else:
            if direction == "min-bound":
                # BV pass condition: fit >= ds * (1 - tol/100)
                ratio = fit_val / ds_val
                ok = ratio >= (1.0 - tol_pct / 100.0)
                delta_pct = (1.0 - ratio) * 100  # negative margin if over-spec
                status = "PASS" if ok else "FAIL"
            else:
                # two-sided: |fit - ds| / ds <= tol/100
                delta_pct = abs(fit_val - ds_val) / ds_val * 100
                status = "PASS" if delta_pct <= tol_pct else "FAIL"
            ws.cell(row=row, column=4, value=_fmt_eng(fit_val)).font = NORMAL_FONT
            ws.cell(row=row, column=5,
                    value=f"{delta_pct:.2f}%" if status != "n/a" else "—").font = NORMAL_FONT
            if status == "PASS":
                n_pass += 1
            else:
                n_fail += 1

        ws.cell(row=row, column=6, value=f"\u00b1{tol_pct:.1f}%").font = NORMAL_FONT
        cell = ws.cell(row=row, column=7, value=status)
        if status == "PASS":
            cell.fill = GOOD_FILL; cell.font = GOOD_FONT
        elif status == "FAIL":
            cell.fill = BAD_FILL; cell.font = BAD_FONT
        elif status == "n/a":
            cell.fill = WARN_FILL; cell.font = WARN_FONT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Overall").font = SECTION_FONT
    ws.cell(row=row, column=4, value=f"{n_pass} PASS  ·  {n_fail} FAIL  ·  {n_na} n/a").font = NORMAL_FONT

    _auto_column_width(ws, min_width=14, max_width=44)


def _build_generic_curve_sheet(ws: Worksheet, *,
                               title: str,
                               x_label: str, y_label: str,
                               ivar: Sequence[float], dvar: Sequence[float],
                               fit: Optional[Sequence[float]],
                               units: str = "") -> None:
    row = 1
    row = _set_title(ws, row, title, cols=6)
    row = _set_section(ws, row,
                       f"Test conditions: {x_label} swept, {y_label} measured"
                       + (f" ({units})" if units else ""), cols=6)

    headers = ["#", f"{x_label}", f"{y_label} (meas)",
               f"{y_label} (fit)" if fit is not None else f"{y_label} (fit)",
               "log10 meas", "log10 fit"]
    row = _set_table_header(ws, row, headers)

    n = len(ivar)
    has_fit = fit is not None and len(fit) == n
    for i in range(n):
        x = float(ivar[i])
        m = float(dvar[i])
        log_m = np.log10(m) if m > 0 else None
        if has_fit:
            f_val = float(fit[i])
            log_f = np.log10(f_val) if f_val > 0 else None
        else:
            f_val = None
            log_f = None
        ws.cell(row=row, column=1, value=i + 1).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=round(x, 4)).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=_fmt_eng(m)).font = NORMAL_FONT
        if has_fit:
            ws.cell(row=row, column=4, value=_fmt_eng(f_val)).font = NORMAL_FONT
        else:
            ws.cell(row=row, column=4, value="n/a").font = NORMAL_FONT
        ws.cell(row=row, column=5,
                value=round(log_m, 4) if log_m is not None else "n/a").font = NORMAL_FONT
        if log_f is not None:
            ws.cell(row=row, column=6, value=round(log_f, 4)).font = NORMAL_FONT
        else:
            ws.cell(row=row, column=6, value="n/a").font = NORMAL_FONT
        row += 1

    if has_fit:
        rms, r2 = _rms_r2(dvar, fit)  # type: ignore[arg-type]
        row += 1
        ws.cell(row=row, column=1, value="Summary").font = SECTION_FONT
        ws.cell(row=row, column=2, value=f"RMS={rms:.4f}").font = NORMAL_FONT
        r2c = ws.cell(row=row, column=3, value=f"R²={r2:.4f}")
        r2_fill, r2_font = _r2_cell_color(r2)
        r2c.font = r2_font
        r2c.fill = r2_fill

    _auto_column_width(ws, min_width=12, max_width=22)


# ----------------------------- public entry -----------------------------

def build_report(*,
                 out_path: Path,
                 device_info: Dict[str, Any],
                 key_params: Dict[str, Any],
                 fit_result: Dict[str, Any],
                 curve_counts: Dict[str, int],
                 curves: Dict[str, Dict[str, Any]],
                 lib_path: Optional[Path] = None) -> None:
    """Write a multi-sheet Excel report.

    Args:
        out_path: where to save the .xlsx file.
        device_info: dict with keys part_number, package, bvdss_v,
                     id_rated_a, vth_typ_v, rdson_max_mohm.
        key_params: dict with datasheet key params, see Summary sheet code.
        fit_result: the dict from task.result after a fit (Total RMS,
                    total r_squared, stages list).
        curve_counts: {curve_name: int} for the Summary sheet.
        curves: dict mapping curve sheet name -> CurveResponse-shaped dict
                with fields "ivar", "dvar", "fit" (optional).
        lib_path: optional path to the exported .lib.  When present, the
                  report includes a Spec_Compliance sheet that reads back
                  VTH0 / RD / RS / BV / RG from the file and compares them
                  to key_params.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1: Summary
    ws = wb.create_sheet("Summary")
    _build_summary_sheet(ws, device_info, key_params, fit_result, curve_counts)

    # Sheet 2: Spec compliance (sits right after Summary so reviewers
    # see the d ata sheet audit on page 2 of the report)
    if lib_path is not None:
        ws_spec = wb.create_sheet("Spec_Compliance")
        _build_spec_compliance_sheet(
            ws_spec, key_params=key_params, lib_path=lib_path,
            bvdss_rated_v=device_info.get("bvdss_v"),
        )

    # Sheet 3..N: curves
    sheet_specs = [
        ("Id-Vg @ Vds=5V",   "idvg_5v"),
        ("Id-Vg @ Vds=0.5V", "idvg_05v"),
        ("Id-Vd",            "idvd"),
        ("Ciss vs Vds",       "cv_ciss"),
        ("Coss vs Vds",       "cv_coss"),
        ("Crss vs Vds",       "cv_crss"),
        ("Body Diode Is-Vsd", "body_diode"),
    ]
    for sheet_name, curve_key in sheet_specs:
        data = curves.get(curve_key, {})
        if not data:
            continue
        ws = wb.create_sheet(sheet_name)
        ivar = data.get("ivar", []) or []
        dvar = data.get("dvar", []) or []
        fit_ = data.get("fit")
        if curve_key == "idvg_5v":
            conditions = "T=25°C, Vds=5V, Vgs swept full Id-Vg range"
            _build_idvg_sheet(ws, title=sheet_name, conditions=conditions,
                               ivar=ivar, dvar=dvar, fit=fit_)
        elif curve_key == "idvg_05v":
            conditions = "T=25°C, Vds=0.5V (linear region extraction); subthreshold mask: Vgs < Vth-0.5"
            _build_idvg_sheet(ws, title=sheet_name, conditions=conditions,
                               ivar=ivar, dvar=dvar, fit=fit_)
        elif curve_key == "idvd":
            conditions = "T=25°C, Id-Vd sweep across multiple Vgs levels"
            _build_idvd_sheet(ws, ivar=ivar, dvar=dvar, fit=fit_,
                              vgs_levels=[5.0, 6.0, 8.0, 10.0])
        elif curve_key == "cv_ciss":
            _build_generic_curve_sheet(
                ws, title=sheet_name, x_label="Vds (V)", y_label="Ciss",
                units="pF", ivar=ivar, dvar=dvar, fit=fit_,
            )
        elif curve_key == "cv_coss":
            _build_generic_curve_sheet(
                ws, title=sheet_name, x_label="Vds (V)", y_label="Coss",
                units="pF", ivar=ivar, dvar=dvar, fit=fit_,
            )
        elif curve_key == "cv_crss":
            _build_generic_curve_sheet(
                ws, title=sheet_name, x_label="Vds (V)", y_label="Crss",
                units="pF", ivar=ivar, dvar=dvar, fit=fit_,
            )
        elif curve_key == "body_diode":
            _build_generic_curve_sheet(
                ws, title=sheet_name, x_label="Vsd (V)", y_label="Is",
                units="A", ivar=ivar, dvar=dvar, fit=fit_,
            )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
