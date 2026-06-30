"""
loader_sdh.py
=============
加载 Silicon-Magic 风格清洗后的 SPICE 数据 Excel。

输入: SDH10N2P1WC-AA_SPICE_Data.xlsx  (用户清洗过的格式)

包含 8 个 sheet:
  - DeviceInfo            器件元数据 (Part Number, Package, BVDSS, ...)
  - ID-VGS_VDS5V          Id-Vg @ Vds=5V (3 列: VGS, ID_25C, ID_150C)
  - ID-VGS_VDS0.5V        Id-Vg @ Vds=0.5V (3 列: VGS, ID_25C, ID_150C)
  - ID-VDS                Id-Vd 多 Vgs (20 列: VDS/ID @ Vgs=5/5.5/6/.../10V)
  - Capacitance_VDS       C-V 曲线 (4 列: VDS, Ciss, Coss, Crss)
  - BodyDiode_IS-VSD      体二极管 (6 列: VSD/IS @ 3 温度)
  - Breakdown_BV          击穿电压
  - SPICE_Params          关键 SPICE 参数 (45 行, 推 BSIM3 初值用)
"""
from __future__ import annotations
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional
import warnings
from openpyxl import load_workbook

warnings.filterwarnings('ignore')


@dataclass
class DeviceInfo:
    """器件元数据"""
    part_number: str = ""
    package: str = ""
    wafer_lot: str = ""
    test_date: str = ""
    test_temp: str = ""
    bvdss_rated_v: float = 0.0
    rdson_max_ohm: float = 0.0     # SI Ω（原始 2.1mΩ → 2.1e-3）
    id_rated_a: float = 0.0
    vth_typ_v: float = 0.0
    extra: dict = field(default_factory=dict)

    def __repr__(self):
        return f"DeviceInfo({self.part_number}, BV={self.bvdss_rated_v}V, RDSon={self.rdson_max_ohm*1e3:.2f}mΩ)"


@dataclass
class SpiceKeyParams:
    """datasheet 关键参数（用于推 BSIM3 初始值）"""
    # Threshold
    vth_25c_v: float = 3.0
    dvth_dT_mv_per_c: float = -9.0  # 典型值

    # On-Resistance (SI Ω)
    rdson_25c_10v_ohm: float = 1.85e-3   # 1.85 mΩ
    rdson_25c_6v_ohm: float = 2.4e-3     # 2.4 mΩ
    rdson_150c_10v_ohm: float = 3.9e-3   # 3.9 mΩ
    rdson_temp_coeff: float = 1.86  # ratio @150/25

    # Transconductance
    gfs_25c_s: float = 250.0

    # Gate Charge
    qg_on_20v_nc: float = 154.0
    qg_on_50v_nc: float = 158.0
    qgs_nc: float = 67.0
    qgd_nc: float = 16.0
    vgs_plateau_v: float = 4.9

    # Capacitance @ Vds=25V
    ciss_25v_pf: float = 13000.0
    coss_25v_pf: float = 4700.0
    crss_25v_pf: float = 174.0

    # Body Diode
    vsd_25c_v: float = 0.9
    vsd_150c_v: float = 0.79

    # Thermal
    rthjc_c_per_w: float = 0.4
    rthja_c_per_w: float = 40.0

    # Gate R
    rg_internal_ohm: float = 1.6

    # Abs Max
    vds_max_v: float = 100.0
    vgs_max_v: float = 20.0
    id_cont_a: float = 100.0
    id_peak_a: float = 400.0
    tj_max_c: float = 150.0

    # Breakdown
    bvdss_0vgs_v: float = 109.0
    bvgss_pos_v: float = 37.0
    bvgss_neg_v: float = 36.6


def _to_float(v) -> Optional[float]:
    """Parse Excel cell into float, tolerant of unit suffixes.

    Handles:
      - Empty / error sentinels (NaN, N/A, #N/A, #REF!, em-dash)
      - SI prefixes (m-Omega / mOhm -> e-3, u-Omega / uOhm -> e-6)
      - Charge / capacitance units (nC -> e-9, pF -> e-12)
      - Trailing temperature (\u00b0C) and units (V/A/W/S/Omega)
      - Leading +/- / ~ error prefix in any supported encoding
    """
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    # Empty / error sentinels
    if s in ("", "NaN", "N/A", "#N/A", "#REF!", "\u2014", "-"):
        return None
    # Strip leading +/- / ~ error prefix
    PREFIX_CHARS = ("\xa1\xa1", "\xa1", "\u00b1", "\u00a1", "~")
    while s and any(s.startswith(p) for p in PREFIX_CHARS):
        for p in PREFIX_CHARS:
            if s.startswith(p):
                s = s[len(p):].lstrip()
                break
    # SI prefix -> scientific notation
    s = re.sub(r"(?i)\bm(\u2126|Ohm)\b", "e-3", s)
    s = re.sub(r"(?i)\bu(\u2126|Ohm)\b", "e-6", s)
    s = re.sub(r"(?i)\bnC\b",             "e-9",  s)
    s = re.sub(r"(?i)\bpF\b",             "e-12", s)
    # Trailing temperature / units
    s = re.sub(r"\u00b0C\s*$", "", s)
    s = re.sub(r"[VAWS\u2126]+\s*$", "", s)
    s = s.strip()
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def load_device_info(wb) -> DeviceInfo:
    ws = wb['DeviceInfo']
    info = DeviceInfo()
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        key = str(row[0]).strip()
        val = row[1] if len(row) > 1 else None
        if key == 'Part Number':
            info.part_number = str(val) if val else ""
        elif key == 'Package':
            info.package = str(val) if val else ""
        elif key == 'Wafer Lot':
            info.wafer_lot = str(val) if val else ""
        elif key == 'Test Date':
            info.test_date = str(val) if val else ""
        elif key == 'Test Temp (RT)':
            info.test_temp = str(val) if val else ""
        elif key == 'BVDSS Rated':
            info.bvdss_rated_v = _to_float(val) or 0.0
        elif key == 'RDSon max':
            info.rdson_max_ohm = _to_float(val) or 0.0
        elif key == 'ID Rated':
            info.id_rated_a = _to_float(val) or 0.0
        elif key == 'VGS(th) typ':
            info.vth_typ_v = _to_float(val) or 0.0
        else:
            if val is not None:
                info.extra[key] = val
    return info


def load_spice_params(wb) -> SpiceKeyParams:
    """从 SPICE_Params sheet 读取 45 个关键参数"""
    ws = wb['SPICE_Params']
    p = SpiceKeyParams()
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 3 or row[0] is None or row[1] is None:
            continue
        cat = str(row[0]).strip() if row[0] else ""
        param = str(row[1]).strip()
        val = _to_float(row[2]) if len(row) > 2 else None
        if val is None:
            continue
        # Threshold
        if param.startswith('VGS(th)') and '25' in param:
            p.vth_25c_v = val
        elif param == 'dVth/dT':
            p.dvth_dT_mv_per_c = val
        # On-Resistance (SPICE_Params 里的数值是 mΩ，需 × 1e-3 转 SI Ω)
        elif 'RDSon @10Vgs, 25' in param:
            p.rdson_25c_10v_ohm = val * 1e-3
        elif 'RDSon @6Vgs, 25' in param:
            p.rdson_25c_6v_ohm = val * 1e-3
        elif 'RDSon @10Vgs, 150' in param:
            p.rdson_150c_10v_ohm = val * 1e-3
        elif 'RDSon temp coeff' in param:
            p.rdson_temp_coeff = val
        # Transconductance
        elif 'gfs' in param and '25' in param:
            p.gfs_25c_s = val
        # Gate Charge
        elif 'Qg(on) @20V' in param:
            p.qg_on_20v_nc = val
        elif 'Qg(on) @50V' in param:
            p.qg_on_50v_nc = val
        elif param == 'Qgs':
            p.qgs_nc = val
        elif param == 'Qgd':
            p.qgd_nc = val
        elif 'Vgs(pl)' in param:
            p.vgs_plateau_v = val
        # Capacitance
        elif param == 'Ciss @25V':
            p.ciss_25v_pf = val
        elif param == 'Coss @25V':
            p.coss_25v_pf = val
        elif param == 'Crss @25V':
            p.crss_25v_pf = val
        # Body Diode
        elif 'VSD @25' in param:
            p.vsd_25c_v = val
        elif 'VSD @150' in param:
            p.vsd_150c_v = val
        # Thermal
        elif 'R��JC' in param or 'RθJC' in param:
            p.rthjc_c_per_w = val
        elif 'R��JA' in param or 'RθJA' in param:
            p.rthja_c_per_w = val
        # Gate R
        elif 'Rg' in param and 'internal' in param:
            p.rg_internal_ohm = val
        # Abs Max
        elif param == 'VDS_max':
            p.vds_max_v = val
        elif param == 'VGS_max':
            p.vgs_max_v = abs(val)
        elif param == 'ID_continuous':
            p.id_cont_a = val
        elif param == 'ID_peak':
            p.id_peak_a = val
        elif param == 'TJ_max':
            p.tj_max_c = val
        # Breakdown
        elif 'BVDSS @0VGS' in param:
            p.bvdss_0vgs_v = val
        elif param == 'BVGSS+':
            p.bvgss_pos_v = val
        elif param == 'BVGSS-':
            p.bvgss_neg_v = val
    return p


def load_idvg(ws, vds_value: float) -> dict:
    """加载 Id-Vg 曲线，返回 {vgs_v, id_a, temperature_c} 列表

    列格式: VGS (V), ID_25C (A), ID_150C (A)
    """
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return {'vds_v': vds_value, 'curves': []}
    header = [str(c) if c else '' for c in rows[0]]
    # 找温度列
    temp_cols = []
    for c, h in enumerate(header):
        if c == 0:
            continue
        if 'ID' in h:
            t = 25
            if '150' in h:
                t = 150
            elif '25' in h:
                t = 25
            elif '-55' in h or 'neg55' in h.lower():
                t = -55
            temp_cols.append((c, t))
    curves = {t: [] for _, t in temp_cols}
    for row in rows[2:]:
        if not row or row[0] is None:
            continue
        vgs = _to_float(row[0])
        if vgs is None:
            continue
        for c, t in temp_cols:
            if c < len(row):
                id_a = _to_float(row[c])
                if id_a is not None:
                    curves[t].append({'vgs_v': vgs, 'id_a': id_a, 'vds_v': vds_value, 'temperature_c': t})
    return {'vds_v': vds_value, 'curves': [curves[t] for t in dict.fromkeys(t for _, t in temp_cols)]}


def load_idvd(ws) -> list:
    """加载 Id-Vd 曲线

    列格式（成对出现）: VDS @VGS=X.XV, ID @VGS=X.XV, ...
    """
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return []
    header = [str(c) if c else '' for c in rows[0]]
    # 找 Vgs 配对
    import re
    pairs = []
    c = 0
    while c < len(header) - 1:
        h0 = header[c]
        h1 = header[c + 1]
        m0 = re.search(r'@?VGS\s*=?\s*([\d.]+)\s*V', h0)
        m1 = re.search(r'@?VGS\s*=?\s*([\d.]+)\s*V', h1)
        if m0 and m1 and float(m0.group(1)) == float(m1.group(1)) and 'VDS' in h0.upper() and 'ID' in h1.upper():
            pairs.append((c, c + 1, float(m0.group(1))))
            c += 2
        else:
            c += 1
    curves = []
    for r in rows[2:]:
        if not r or all(v is None for v in r):
            continue
        for vds_col, id_col, vgs_v in pairs:
            if vds_col >= len(r) or id_col >= len(r):
                continue
            vds = _to_float(r[vds_col])
            id_a = _to_float(r[id_col])
            if vds is None or id_a is None:
                continue
            if vds <= 0:
                continue
            curves.append({'vds_v': vds, 'id_a': id_a, 'vgs_v': vgs_v, 'temperature_c': 25})
    return curves


def load_cv_vds(ws) -> list:
    """加载 C-V 曲线

    列格式: VDS (V), Ciss (pF), Coss (pF), Crss (pF)
    """
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return []
    curves = []
    for r in rows[2:]:
        if not r or r[0] is None:
            continue
        vds = _to_float(r[0])
        ciss = _to_float(r[1]) if len(r) > 1 else None
        oss = _to_float(r[2]) if len(r) > 2 else None
        rss = _to_float(r[3]) if len(r) > 3 else None
        if vds is None:
            continue
        curves.append({
            'vds_v': vds,
            'ciss_pf': ciss,
            'coss_pf': oss,
            'crss_pf': rss,
        })
    return curves


def load_bodydiode(ws) -> list:
    """加载体二极管

    列格式: VSD_-55C, IS_-55C, VSD_25C, IS_25C, VSD_150C, IS_150C
    """
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return []
    header = [str(c) if c else '' for c in rows[0]]
    pairs = []
    for c, h in enumerate(header):
        if 'VSD' in h.upper():
            t = 25
            if '-55' in h or 'neg55' in h.lower():
                t = -55
            elif '150' in h:
                t = 150
            if c + 1 < len(header) and 'IS' in header[c + 1].upper():
                pairs.append((c, c + 1, t))
    curves = []
    for r in rows[2:]:
        if not r:
            continue
        for vsd_col, is_col, t in pairs:
            if vsd_col >= len(r) or is_col >= len(r):
                continue
            vsd = _to_float(r[vsd_col])
            is_a = _to_float(r[is_col])
            if vsd is None or is_a is None:
                continue
            curves.append({'vsd_v': abs(vsd), 'is_a': abs(is_a), 'temperature_c': t, 'vgs_v': 0})
    return curves


@dataclass
class SpiceDataSet:
    """完整的 SPICE 数据集"""
    device_info: DeviceInfo
    key_params: SpiceKeyParams
    idvg_vds5: list      # [{vgs_v, id_a, vds_v, temperature_c}, ...]
    idvg_vds05: list     # 同上
    idvd: list           # [{vds_v, id_a, vgs_v, temperature_c}, ...]
    cv_vds: list         # [{vds_v, ciss_pf, coss_pf, crss_pf}, ...]
    body_diode: list     # [{vsd_v, is_a, temperature_c, vgs_v}, ...]

    def summary(self) -> str:
        lines = [
            f"Device: {self.device_info.part_number}",
            f"  Package: {self.device_info.package}",
            f"  BVdss: {self.device_info.bvdss_rated_v}V, RDSon: {self.device_info.rdson_max_ohm*1e3:.2f}mΩ",
            f"  Vth(typ): {self.device_info.vth_typ_v}V",
            f"Curves:",
            f"  Id-Vg @5V:    {len(self.idvg_vds5)} points",
            f"  Id-Vg @0.5V:  {len(self.idvg_vds05)} points",
            f"  Id-Vd:        {len(self.idvd)} points",
            f"  C-V:          {len(self.cv_vds)} points",
            f"  Body Diode:   {len(self.body_diode)} points",
            f"Key SPICE params:",
            f"  Vth={self.key_params.vth_25c_v}V, dVth/dT={self.key_params.dvth_dT_mv_per_c}mV/°C",
            f"  RDSon@25C,10V={self.key_params.rdson_25c_10v_ohm*1e3:.2f}mΩ, @150C={self.key_params.rdson_150c_10v_ohm*1e3:.2f}mΩ",
            f"  Ciss/Coss/Crss@25V = {self.key_params.ciss_25v_pf:.0f}/{self.key_params.coss_25v_pf:.0f}/{self.key_params.crss_25v_pf:.0f} pF",
            f"  Qg@20V={self.key_params.qg_on_20v_nc}nC, Qgd={self.key_params.qgd_nc}nC, Vgs_pl={self.key_params.vgs_plateau_v}V",
        ]
        return '\n'.join(lines)


def load_sdh_excel(filepath: str | Path) -> SpiceDataSet:
    """加载 SDH 风格的 Excel 数据"""
    fp = Path(filepath)
    if not fp.exists():
        raise FileNotFoundError(f"File not found: {fp}")
    wb = load_workbook(fp, data_only=True)

    info = load_device_info(wb)
    params = load_spice_params(wb)

    idvg5 = load_idvg(wb['ID-VGS_VDS5V'], vds_value=5.0)['curves'][0] if 'ID-VGS_VDS5V' in wb.sheetnames else []
    idvg05 = load_idvg(wb['ID-VGS_VDS0.5V'], vds_value=0.5)['curves'][0] if 'ID-VGS_VDS0.5V' in wb.sheetnames else []
    idvd = load_idvd(wb['ID-VDS']) if 'ID-VDS' in wb.sheetnames else []
    cv = load_cv_vds(wb['Capacitance_VDS']) if 'Capacitance_VDS' in wb.sheetnames else []
    body = load_bodydiode(wb['BodyDiode_IS-VSD']) if 'BodyDiode_IS-VSD' in wb.sheetnames else []

    return SpiceDataSet(
        device_info=info,
        key_params=params,
        idvg_vds5=idvg5,
        idvg_vds05=idvg05,
        idvd=idvd,
        cv_vds=cv,
        body_diode=body,
    )


if __name__ == '__main__':
    ds = load_sdh_excel('datademo/SDH10N2P1WC-AA_SPICE_Data.xlsx')
    print(ds.summary())
